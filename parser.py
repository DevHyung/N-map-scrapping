import json
import requests
from selenium import webdriver
from urllib.parse   import quote
from CONFIG import *
from bs4 import BeautifulSoup
import time
import random
from math import ceil
from openpyxl import load_workbook,Workbook
import os
def get_url_encode(_query):
    return quote(_query).replace('%20', '+')

def get_cookie():
    driver.find_element_by_xpath('//*[@id="search-input"]').clear()
    driver.find_element_by_xpath('//*[@id="search-input"]').send_keys(keyword + '\n')
    time.sleep(2)
    # header setting
    return get_cookie_str(driver.get_cookies())

def get_cookie_str(_cookie):
    cookieStr = ""
    for cookie in _cookie:
        cookieStr += "{}={};".format(cookie['name'],cookie['value'])
    return  cookieStr

def get_info():
    depth2List = []
    for pageIdx in range(1,lastPage+1):
        log('s',"{} 페이지 추출중...".format(pageIdx))
        html = requests.get(url.format(get_url_encode(keyword), pageIdx), headers=headers)
        jsonStr = json.loads(html.text)
        try:
            results = jsonStr['result']['site']['list']
        except:
            log('e', '쿠키 유효기간 만기로 재갱신중')
            headers['cookie'] = get_cookie()
            log('s', "{} 페이지 추출중...".format(pageIdx))
            html = requests.get(url.format(get_url_encode(keyword), pageIdx), headers=headers)
            jsonStr = json.loads(html.text)
            results = jsonStr['result']['site']['list']
        if len(results) != 0:
            for result in results:
                name = result['name']
                address = result['address']
                roadAddress = result['roadAddress']
                tel = result['tel']
                code = result['id'][1:]
                homepage = result['homePage']
                depth2List.append([keyword,name, roadAddress, address, tel, homepage, detailBaseUrl + code])
        else:
            log('s',"끝")

        if len(depth2List) >= 50:
            save_excel(FILENAME, depth2List, None)  # init
            depth2List.clear()
        time.sleep(random.randint(MIN,MAX))
    save_excel(FILENAME, depth2List, None)  # init

def log(tag, text):
	# Info tag
	if(tag == 'i'):
		print("[INFO] " + text)
	# Error tag
	elif(tag == 'e'):
		print("[ERROR] " + text)
	# Success tag
	elif(tag == 's'):
		print("[SUCCESS] " + text)

def README():
    print("_____ Copyright @Hyung June, Park ____________________________________")
    print("# 사용법 :")
    print("# 1. 저장 파일의 이름만 적는다 (기본 포맷은 xlsx 이며 확장자는 적지 않는다 ) ")
    print("#     1-1. 위에 적은 파일은 프로그램이 종료전까지 계속 데이터가 쌓인다")
    print("# 2. 페이지 -> 페이지 파싱간의 딜레이 최소, 최대 정수값을 입력한다. ")
    print("#     2-1. 예를들어 3,10을 입력하면 페이지간 딜레이가 3~10사이의 랜덤정수가 된다.")
    print("# 3. 검색할 키워드를 입력 한다. ")
    print("#     3-1. 끝내시려면 숫자 0을 입력하시고 엔터를 누르시면 됩니다.")
    print("#_____________________________________________________________________")

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for depth1List in _DATA:
            sheet.append(depth1List)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        book.save(_FILENAME)

if __name__ == "__main__":
    # CODE HERE
    # driver setting
    README()
    driver = webdriver.Chrome('./chromedriver')
    driver.maximize_window()
    driver.get('https://map.naver.com/')
    # get cookie
    FILENAME = input(">>> 저장 파일 *이름만* 적어주세요 ( 확장자 미포함 ) :") + '.xlsx'
    MIN = int ( input(">>> 페이지 딜레이 최소 정수값 : ") )
    MAX = int ( input(">>> 페이지 딜레이 최대 정수값 : ") )
    HEADER = ['키워드','업소명', '도로명주소', '지번주소', '전화번호', '홈페이지1', '상세페이지']
    save_excel(FILENAME,None,HEADER) # init

    while True:
        keyword = input(">>> 검색하실 키워드를 입력(끝내려면 0 입력) : ").strip()
        if keyword == '0':
            break
        headers['cookie'] = get_cookie()

        # get total cnt
        bs4 = BeautifulSoup(driver.page_source,'lxml')
        totalCnt = int ( bs4.find('span',class_='n').em.get_text().strip().replace(',','') )
        lastPage = ceil(totalCnt/10)

        log('i',"총 {}개 검색결과 존재".format(totalCnt))
        log('i',"마지막 페이지 {}".format(lastPage))

        # parsing
        get_info()
        print('#_____________________________________________________________________')
        # End
    driver.quit()

